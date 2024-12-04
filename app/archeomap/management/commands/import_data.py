# archeomap/management/commands/import_data.py
import os
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from archeomap.models.models import Monuments

class Command(BaseCommand):
    help = 'Импортирует данные памятников из Excel файла в базу данных'

    def handle(self, *args, **kwargs):
        # Указываем путь к файлу Excel
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__name__)))
        file_path = os.path.join(base_dir, 'app', 'Data.xlsx')

        try:
            # Загружаем рабочую книгу Excel
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Проверяем заголовки столбцов
            headers = [cell.value for cell in sheet[1]]
            required_columns = ['TITLE WITH ID', 'TITLE FOR USERS', 'DESCRIPTION', 
                                'LATITUDE', 'LONGITUDE', 'LANDMARK', 'ADDRESS']
            missing_columns = [col for col in required_columns if col not in headers]

            if missing_columns:
                self.stdout.write(self.style.ERROR(f'Отсутствуют обязательные колонки: {", ".join(missing_columns)}'))
                return

            # Индексы нужных колонок
            column_indices = {col: headers.index(col) + 1 for col in required_columns}

            # Проходим по строкам и создаем/обновляем объекты в базе данных
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    title = row[column_indices['TITLE WITH ID'] - 1]
                    name = row[column_indices['TITLE FOR USERS'] - 1]
                    description = row[column_indices['DESCRIPTION'] - 1]
                    latitude = row[column_indices['LATITUDE'] - 1]
                    longitude = row[column_indices['LONGITUDE'] - 1]
                    landmark = row[column_indices['LANDMARK'] - 1]
                    address = row[column_indices['ADDRESS'] - 1]

                    if title and name:  # Проверяем, что обязательные поля заполнены
                        # Проверяем, существует ли запись с таким title или name
                        if Monuments.objects.filter(title=title).exists():
                            self.stdout.write(self.style.WARNING(f'Пропущен памятник с title: {title}, так как он уже существует'))
                            continue
                        if Monuments.objects.filter(name=name).exists():
                            self.stdout.write(self.style.WARNING(f'Пропущен памятник с name: {name}, так как он уже существует'))
                            continue

                        # Если запись не найдена, создаём её
                        monument = Monuments.objects.create(
                            title=title,
                            name=name,
                            description=description,
                            latitude=latitude,
                            longitude=longitude,
                            landmark=landmark,
                            address=address,
                        )
                        self.stdout.write(self.style.SUCCESS(f'Создан памятник: {monument.name}'))

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'Ошибка при обработке строки {row}: {e}'))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Ошибка при открытии файла: {e}'))
