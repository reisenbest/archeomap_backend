'''
1. список полей из которых инфа должна грузиться перечисляем в списке

2. проверка на существование. по первым 4 цифрам. если совпадают, то прпоускать? обновлять? 
если ЕСТЬ такой код то создание (функция обновления объекта)
если такого кода НЕТ то создание (функция создания объекта)

3. 



'''


from django.core.management.base import BaseCommand
from archeomap.models.models import Monuments, Dating, Classification, CustomCategory, ResearchYears, ExcavationsSquare


from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path



class Command(BaseCommand):
    help = "Импорт данных из Excel файла в базу данных"

    col_names_with_indexes ={
                            'ID': 0,
                              'TITLE WITH ID': 1,
                                'TITLE FOR USERS': 2,
                                  'DESCRIPTION': 3,
                                    'LANDMARK': 4,
                                      'ADDRESS': 5,
                                        'DATING': 6,
                                          'CLASSIFICATION CATEGORY': 7,
                                            'CUSTOM CATEGORY': 8,
                                              'LATITUDE': 9,
                                                'LONGITUDE': 10,
                                                  'RESEARCH YEARS': 11,
                                                    'ORGANIZATIONS': 12,
                                                      'AUTHORS': 13,
                                                        'BIBLIOGRAPHY': 14,
                                                          'ADDITIONAL CONTENT': 15,
                                                            'SLUG': 16,
                                                              'VISIBLE': 17,
                                                                'EXCAVATIONS SQUARE': 18
                                                              }


    def add_arguments(self, parser):
        parser.add_argument(
            'filepath',
            type=Path,
            help='./Data.xlsx'
        )

    def handle(self, *args, **options):
        filepath: Path = options['filepath']

        # Нормализуем (преобразуем относительный путь в абсолютный)
        filepath = filepath.expanduser().resolve()

        # Проверки
        if not filepath.exists():
            raise CommandError(f"Файл не найден: {filepath}")
        if not filepath.is_file():
            raise CommandError(f"Путь не является файлом: {filepath}")
        if filepath.suffix.lower() not in ('.xlsx', '.xlsm', '.xltx'):
            raise CommandError(f"Похоже, это не Excel файл (ожидается .xlsx/.xlsm/.xltx): {filepath}")

        self.stdout.write(self.style.SUCCESS(f"Открываю файл: {filepath}"))
        # дальше — чтение openpyxl: load_workbook(filename=str(filepath)) и т.д.p

        # === ЧТЕНИЕ EXCEL ===
        workbook = load_workbook(filename=str(filepath))
        sheet = workbook.active  # первый лист

        rows_count = self.count_non_empty_rows(sheet)

        for i in range(2, rows_count + 1):  # начинаем с 2, чтобы пропустить заголовки
          row = sheet[i]  # строка Excel

          name_value = row[self.col_names_with_indexes['TITLE FOR USERS']].value
          title_value = row[self.col_names_with_indexes['TITLE WITH ID']].value
          latitude_value = row[self.col_names_with_indexes['LATITUDE']].value
          longitude_value = row[self.col_names_with_indexes['LONGITUDE']].value

          # Проверка обязательных полей
          if not (name_value and title_value and latitude_value and longitude_value):
              self.stdout.write(self.style.WARNING(f"Пропущена строка {i} — отсутствуют обязательные данные"))
              continue

          # Преобразуем LATITUDE и LONGITUDE в float, заменяя запятую на точку
          try:
              latitude_value = float(str(latitude_value).replace(',', '.'))
              longitude_value = float(str(longitude_value).replace(',', '.'))
          except ValueError:
              self.stdout.write(self.style.WARNING(f"Ошибка конвертации координат в строке {i}"))
              continue

          # Ищем существующий объект по title
          monument, created = Monuments.objects.update_or_create(
              title=title_value,
              defaults={
                  'name': name_value,
                  'latitude': latitude_value,
                  'longitude': longitude_value,
              }
          )

          if created:
              self.stdout.write(self.style.SUCCESS(f"Добавлен памятник: {name_value}"))
          else:
              self.stdout.write(self.style.SUCCESS(f"Обновлён памятник: {name_value}"))


      

    def count_non_empty_rows(self, sheet: Worksheet) -> int:
      """
      Возвращает количество строк в листе, где хотя бы одна ячейка не пустая.
      """
      count = 0
      for row in sheet.iter_rows(values_only=True):
          if any(cell is not None for cell in row):
              count += 1
      return count


