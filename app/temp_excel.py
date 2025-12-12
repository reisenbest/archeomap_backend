import openpyxl

path_to_excel = r"C:\Users\Aleksey\Desktop\Programmatura_computatralis\archeomap_backend\app\Data.xlsx"

workbook = openpyxl.load_workbook(filename=str(path_to_excel))
sheet = workbook.active  # первый лист

# Получаем первую строку — заголовки колонок
header_row = next(sheet.iter_rows(values_only=True))

col_names_with_indexes ={
  'ID': 0,
    'TITLE WITH ID': 1,
      'TITLE FOR USERS': 2,
        'DESCRIPTION': 3,
          'LANDMARK': 4,
            'ADDRESS': 5,
              'DATING': 6,
                'CLASSIFICATION CATEGORY': 7,
                  'CUSTOM CATEGORY': 8,
                    'LATITUDE': 9,
                      'LONGITUDE': 10,
                        'RESEARCH YEARS': 11,
                          'ORGANIZATIONS': 12,
                            'AUTHORS': 13,
                              'BIBLIOGRAPHY': 14,
                                'ADDITIONAL CONTENT': 15,
                                  'SLUG': 16,
                                    'VISIBLE': 17,
                                      'EXCAVATIONS SQUARE': 18
                                    }

for i, col in enumerate(header_row, start=0): 
  print(f"{i}: {col}") 

print('загаловки таблицы')

def count_non_empty_rows(sheet: sheet) -> int:
    """
    Возвращает количество строк в листе, где хотя бы одна ячейка не пустая.
    """
    count = 0
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            count += 1
    return count


rows_count = count_non_empty_rows(sheet)
print(f"Количество непустых строк: {rows_count}")

# values = [cell.value for cell in sheet[3]]
# print(values)
"""
НОМЕР СТРОКИ НАЧИНАЕТСЯ С 1!! НОМЕР ЯЧЕЙКИ С 0!!
"""

for i in range(1, rows_count+1):
    print(sheet[i][2].value)
